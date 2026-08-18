import io

from flask import Blueprint, jsonify, Response
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Evaluation
from routes.courses import _assert_owns_course

dashboard_bp = Blueprint(
    "dashboard", __name__, url_prefix="/courses/<int:course_id>/evaluations/<int:evaluation_id>"
)


def _get_owned_evaluation(course_id, evaluation_id):
    if not _assert_owns_course(course_id):
        return None
    return Evaluation.first(id=evaluation_id, course_id=course_id)


@dashboard_bp.route("/completion", methods=["GET"])
@login_required
def completion(course_id, evaluation_id):
    """Who has and hasn't submitted yet, grouped by group."""
    evaluation = _get_owned_evaluation(course_id, evaluation_id)
    if not evaluation:
        return jsonify({"error": "Evaluation not found"}), 404
    return jsonify(evaluation.completion())


@dashboard_bp.route("/results", methods=["GET"])
@login_required
def results(course_id, evaluation_id):
    """Per-student aggregate averages plus the full individual evaluator breakdown."""
    evaluation = _get_owned_evaluation(course_id, evaluation_id)
    if not evaluation:
        return jsonify({"error": "Evaluation not found"}), 404
    return jsonify(evaluation.results())


@dashboard_bp.route("/results/export.xlsx", methods=["GET"])
@login_required
def export_xlsx(course_id, evaluation_id):
    """
    An Excel workbook with two sheets: per-student averages (same table as
    the Performance Table tab), and a colored evaluator x rated matrix per
    group (same layout and colors as the Matrix tab: yellow diagonal = self,
    red = evaluator didn't submit / didn't rate that person, white = the
    total they gave). This has to be .xlsx rather than .csv — plain CSV has
    no concept of cell color, and the whole point of the matrix export is
    the color coding.
    """
    evaluation = _get_owned_evaluation(course_id, evaluation_id)
    if not evaluation:
        return jsonify({"error": "Evaluation not found"}), 404

    data = evaluation.results()
    completion_groups = evaluation.completion()
    criteria_names = [c["name"] for c in data["criteria"]]

    # Same total-per-pair computation the Matrix tab does client-side in
    # results.js — kept in sync deliberately, not shared code, since one is
    # Python and one is JS; if the matrix logic ever changes, both need it.
    totals_by_pair = {}
    for row in data["individual_scores"]:
        key = (row["evaluator_id"], row["ratee_id"])
        totals_by_pair[key] = totals_by_pair.get(key, 0) + row["score"]

    wb = Workbook()

    # --- Sheet 1: Averages per student ---
    ws = wb.active
    ws.title = "Averages"
    header_fill = PatternFill("solid", fgColor="F1F5F9")  # slate-100, matches the site's table header row
    bold = Font(bold=True)
    thin = Side(style="thin", color="E2E8F0")  # slate-200
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header = ["Student"] + criteria_names + ["Total"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for s in data["aggregates"]:
        by_name = {c["criterion"]: c["average"] for c in s["by_criterion"]}
        row = [s["name"]] + [by_name.get(c, "") for c in criteria_names] + [s["total"]]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            if cell.column > 1:
                cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 26
    for i in range(len(criteria_names)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16
    ws.freeze_panes = "B2"

    # --- Sheet 2: Matrix (one block per group) ---
    ws2 = wb.create_sheet("Matrix")
    yellow_fill = PatternFill("solid", fgColor="FDE047")   # matches .matrix-cell--self
    yellow_font = Font(color="78350F")
    red_fill = PatternFill("solid", fgColor="EF4444")      # matches .matrix-cell--missing
    red_font = Font(color="FFFFFF")
    group_fill = PatternFill("solid", fgColor="F8FAFC")    # matches .matrix-corner / card header
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 1
    for group in completion_groups:
        students = group["students"]
        if not students:
            continue

        width = len(students) + 1
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
        title_cell = ws2.cell(row=r, column=1, value=group["group_label"])
        title_cell.font = Font(bold=True)
        title_cell.fill = group_fill
        title_cell.alignment = Alignment(horizontal="left")
        r += 1

        header_row = r
        ws2.cell(row=header_row, column=1).border = border
        ws2.cell(row=header_row, column=1).fill = group_fill
        for i, evaluator in enumerate(students):
            cell = ws2.cell(row=header_row, column=i + 2, value=evaluator["name"])
            cell.font = bold
            cell.fill = group_fill
            cell.alignment = center
            cell.border = border
        r += 1

        for ratee in students:
            row_cell = ws2.cell(row=r, column=1, value=ratee["name"])
            row_cell.font = bold
            row_cell.fill = group_fill
            row_cell.border = border
            row_cell.alignment = Alignment(horizontal="left", vertical="center")

            for i, evaluator in enumerate(students):
                cell = ws2.cell(row=r, column=i + 2)
                cell.border = border
                cell.alignment = center
                if evaluator["id"] == ratee["id"]:
                    cell.value = "—"
                    cell.fill = yellow_fill
                    cell.font = yellow_font
                elif not evaluator["has_submitted"]:
                    cell.value = "—"
                    cell.fill = red_fill
                    cell.font = red_font
                else:
                    total = totals_by_pair.get((evaluator["id"], ratee["id"]))
                    if total is None:
                        cell.value = "—"
                        cell.fill = red_fill
                        cell.font = red_font
                    else:
                        cell.value = total
            r += 1

        r += 2  # blank rows between group blocks

    ws2.column_dimensions["A"].width = 22
    max_students_in_a_group = max((len(g["students"]) for g in completion_groups), default=0)
    for i in range(max_students_in_a_group):
        ws2.column_dimensions[get_column_letter(i + 2)].width = 16
    ws2.freeze_panes = "B1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "".join(c if c.isalnum() or c in " -_" else "_" for c in evaluation.title).strip() or "results"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
